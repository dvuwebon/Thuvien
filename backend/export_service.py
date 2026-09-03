import io
import json
import csv
from datetime import datetime
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import qrcode
from PIL import Image

def generate_books_excel(books: List[Dict[str, Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách Sách"

    # Header styling
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    headers = ["Mã sách", "Tên sách", "Tác giả", "Thể loại", "Tổng SL", "Đang mượn", "Còn lại", "Mô tả ngắn"]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

    for row_idx, book in enumerate(books, start=2):
        qty = int(book.get("quantity", 1))
        borrowed = int(book.get("borrowed", 0))
        available = max(0, qty - borrowed)
        
        row_data = [
            book.get("id"),
            book.get("title", ""),
            book.get("author", "Chưa rõ"),
            book.get("category", "Khác"),
            qty,
            borrowed,
            available,
            (book.get("desc") or "")[:150]
        ]
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 22
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = cell_font
            cell.border = thin_border
            if col_num in [1, 5, 6, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Adjust column widths
    widths = [10, 35, 25, 22, 12, 14, 12, 45]
    for i, w in enumerate(widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_borrows_excel(records: List[Dict[str, Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lịch sử Mượn Trả"

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    headers = ["Mã phiếu", "Tên sách", "Độc giả", "Hình thức", "Ngày mượn", "Hạn trả", "Ngày trả thực tế", "Trạng thái"]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

    for row_idx, r in enumerate(records, start=2):
        b_date = r.get("borrowDate", "")
        if b_date and len(b_date) >= 10:
            b_date = b_date[:10]
        r_date = r.get("returnDate", "")
        if r_date and len(r_date) >= 10:
            r_date = r_date[:10]
        act_date = r.get("actualReturnDate", "")
        if act_date and len(act_date) >= 10:
            act_date = act_date[:10]
        else:
            act_date = "-"

        row_data = [
            r.get("id"),
            r.get("bookTitle", ""),
            r.get("readerName", ""),
            r.get("borrowType", "Mượn về nhà"),
            b_date,
            r_date,
            act_date,
            r.get("status", "Chờ duyệt")
        ]
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 22
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = cell_font
            cell.border = thin_border
            if col_num in [1, 4, 5, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    widths = [12, 35, 25, 20, 15, 15, 18, 16]
    for i, w in enumerate(widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_readers_csv(readers: List[Dict[str, Any]]) -> str:
    output = io.StringIO()
    # Write UTF-8 BOM so Excel opens with proper Vietnamese diacritics
    output.write('\ufeff')
    writer = csv.writer(output)
    writer.writerow(["ID", "Họ và tên", "Tên đăng nhập", "Email", "Số điện thoại", "Địa chỉ", "Ngày sinh"])
    for r in readers:
        writer.writerow([
            r.get("id"),
            r.get("fullName", ""),
            r.get("username", ""),
            r.get("email", ""),
            r.get("phone", ""),
            r.get("address", ""),
            r.get("birthDate", "")
        ])
    return output.getvalue()


def generate_qr_code(data: str) -> bytes:
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=3,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#1E40AF", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def generate_borrow_receipt_pdf(record: Dict[str, Any], book: Optional[Dict[str, Any]] = None, reader: Optional[Dict[str, Any]] = None) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'ReceiptTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#1E40AF'),
        alignment=1
    )
    subtitle_style = ParagraphStyle(
        'ReceiptSub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=11,
        leading=14,
        textColor=colors.HexColor('#64748B'),
        alignment=1
    )
    body_style = ParagraphStyle(
        'ReceiptBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=15,
        textColor=colors.HexColor('#1E293B')
    )

    story.append(Paragraph("SMARTLIB - HE THONG THU VIEN THONG MINH", title_style))
    story.append(Paragraph("PHIEU XAC NHAN MUON SACH / BORROW RECEIPT", subtitle_style))
    story.append(Spacer(1, 20))

    # Info table
    b_date = (record.get("borrowDate") or "")[:10]
    r_date = (record.get("returnDate") or "")[:10]
    status = record.get("status", "Chờ duyệt")

    data = [
        ["Ma phieu (Receipt ID):", f"#{record.get('id')}"],
        ["Ten sach (Book Title):", record.get("bookTitle", "")],
        ["Doc gia (Reader Name):", record.get("readerName", "")],
        ["Hinh thuc muon (Type):", record.get("borrowType", "Muon ve nha")],
        ["Ngay muon (Borrow Date):", b_date],
        ["Han tra (Due Date):", r_date],
        ["Trang thai (Status):", status]
    ]

    t = Table(data, colWidths=[180, 320])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
        ('TEXTCOLOR', (0,0), (0,-1), colors.HexColor('#475569')),
        ('TEXTCOLOR', (1,0), (1,-1), colors.HexColor('#0F172A')),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME', (1,0), (1,-1), 'Helvetica'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
    ]))
    story.append(t)
    story.append(Spacer(1, 25))

    # Add QR code image
    qr_bytes = generate_qr_code(f"SMARTLIB-RECEIPT-{record.get('id')}-{record.get('bookTitle')}")
    qr_io = io.BytesIO(qr_bytes)
    story.append(RLImage(qr_io, width=120, height=120))
    story.append(Spacer(1, 15))
    story.append(Paragraph("Vui long xuat trinh ma QR tren khi den nhan hoac tra sach tai quay thu vien.", subtitle_style))

    doc.build(story)
    return buf.getvalue()